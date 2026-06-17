import os, sys, re, io, uuid, json, base64, traceback, math, random, csv
from datetime import datetime, timedelta, date
_ML_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _ML_DIR)
try:
    import ml_engine as ML
    print("✅ ML engine loaded")
except ImportError as e:
    print(f"❌ ML engine not found: {e}\n   Place ml_engine.py in the same folder.")
    ML = None
try:
    from flask import Flask, jsonify, request, send_file, send_from_directory, session
    from flask_cors import CORS
except ImportError:
    print("❌  Run: pip install flask flask-cors"); sys.exit(1)
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
app = Flask(__name__, static_folder=".")
app.secret_key = "finup_2026_secret"
CORS(app, origins="*")
UPLOAD_FOLDER = os.path.join(_ML_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
import sqlite3
DATABASE = os.path.join(_ML_DIR, "finup.db")
def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn
def normalize_username(username):
    return re.sub(r"[^a-z0-9_.-]", "", (username or "").strip().lower())
def build_storage_name(user_id, filename):
    ext = os.path.splitext(filename or "")[1]
    return f"{user_id}_{uuid.uuid4().hex}{ext}"
def current_user_id():
    return session.get("user_id", "")
def current_user_profile():
    uid = current_user_id()
    if not uid:
        return None
    return {
        "id": uid,
        "username": session.get("username", ""),
        "display_name": session.get("display_name", ""),
    }
def auth_response_payload():
    user = current_user_profile()
    return {"authenticated": bool(user), "user": user}
def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user_id():
            return jsonify({"error": "Authentication required"}), 401
        return fn(*args, **kwargs)
    return wrapper
def require_report_owner(conn, report_id, user_id):
    return conn.execute(
        "SELECT * FROM reports WHERE id=? AND user_id=?",
        (report_id, user_id)
    ).fetchone()
def clean_upload_name(name):
    return os.path.basename((name or "upload").strip()) or "upload"
def upload_path(name):
    return os.path.join(UPLOAD_FOLDER, clean_upload_name(name))
def save_uploaded_statement(name, raw):
    path = upload_path(name)
    with open(path, "wb") as fh:
        fh.write(raw)
    return path
def get_saved_statement_names():
    return {
        entry.name for entry in os.scandir(UPLOAD_FOLDER)
        if entry.is_file()
    }
def sync_storage(conn, user_id=None):
    try:
        existing_files = get_saved_statement_names()
        user_filter = " WHERE user_id=?" if user_id else ""
        report_file_col = "storage_name"
        if existing_files:
            placeholders = ",".join("?" for _ in existing_files)
            if user_id:
                conn.execute(
                    f"DELETE FROM transactions WHERE user_id=? AND source_file NOT IN ({placeholders})",
                    (user_id, *tuple(existing_files))
                )
                conn.execute(
                    f"DELETE FROM reports WHERE user_id=? AND {report_file_col} NOT IN ({placeholders})",
                    (user_id, *tuple(existing_files))
                )
            else:
                conn.execute(f"DELETE FROM transactions WHERE source_file NOT IN ({placeholders})", tuple(existing_files))
                conn.execute(f"DELETE FROM reports WHERE {report_file_col} NOT IN ({placeholders})", tuple(existing_files))
        else:
            conn.execute(f"DELETE FROM transactions{user_filter}", (user_id,) if user_id else ())
            conn.execute(f"DELETE FROM reports{user_filter}", (user_id,) if user_id else ())
        conn.commit()
    except sqlite3.Error as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        print(f"Storage sync skipped: {exc}")
def _goal_to_dict(row):
    if not row:
        return None
    g = dict(row)
    target = float(g.get("target", 0) or 0)
    saved = float(g.get("saved", 0) or 0)
    g["percent"] = round(saved / target * 100, 1) if target else 0
    return g
def fetch_goal_state(conn, user_id):
    goals = [
        _goal_to_dict(r) for r in conn.execute(
            "SELECT * FROM goals WHERE user_id=? AND completed=0 ORDER BY created_at DESC",
            (user_id,)
        ).fetchall()
    ]
    achieved = [
        dict(r) for r in conn.execute(
            "SELECT * FROM achieved_goals WHERE user_id=? ORDER BY achieved_at DESC",
            (user_id,)
        ).fetchall()
    ]
    return goals, achieved
def fetch_goal_payload(conn, user_id):
    reload_txns(conn, user_id)
    goals, achieved = fetch_goal_state(conn, user_id)
    total_credit = float(DB["summary"].get("total_credit", 0) or 0)
    total_debit = float(DB["summary"].get("total_debit", 0) or 0)
    net = round(total_credit - total_debit, 2)
    rate = round((net / total_credit) * 100, 1) if total_credit > 0 else 0
    return {
        "goals": goals,
        "achieved_goals": achieved,
        "stats": {
            "monthly_savings": net,
            "savings_rate": rate,
            "achieved_count": len(achieved),
            "active_count": len(goals),
            "total_target": round(sum(float(g.get("target", 0) or 0) for g in goals), 2),
            "total_saved": round(sum(float(g.get("saved", 0) or 0) for g in goals), 2),
        }
    }
def _txn_row_to_ml_txn(row):
    amount = float(row.get("amount", 0) or 0)
    signed_amount = amount if row.get("type") == "credit" else -amount
    category = row.get("category", "Others") or "Others"
    return {
        "date": row.get("date", ""),
        "description": row.get("description", ""),
        "amount": round(signed_amount, 2),
        "balance": float(row.get("balance", 0) or 0),
        "mode": row.get("mode", "UPI") or "UPI",
        "category": category,
        "confidence": float(row.get("confidence", 0) or 0),
        "color": CAT_COLORS.get(category, "#6b7280"),
        "icon": "💳",
        "is_card": int(row.get("is_card", 0) or 0),
        "source_file": row.get("source_file", ""),
    }
def build_persisted_ml_result(rows):
    txns = [_txn_row_to_ml_txn(r) for r in rows]
    if not txns:
        return {
            "transactions": [],
            "summary": {},
            "income": 0,
            "spent": 0,
            "net": 0,
            "savings_rate": 0,
            "mode_breakdown": {},
            "anomalies": [],
            "forecast": {},
            "bills": [],
            "investments": {"detected_txns": [], "total_invested": 0, "count": 0},
            "goal_projection": {"goals": [], "monthly_savings": 0, "savings_rate": 0},
            "insights": [],
            "cashflow": [],
            "total_txns": 0,
            "account": {},
        }
    credits = [t for t in txns if t["amount"] > 0]
    debits = [t for t in txns if t["amount"] < 0]
    income = round(sum(t["amount"] for t in credits), 2)
    spent = round(sum(abs(t["amount"]) for t in debits), 2)
    net = round(income - spent, 2)
    rate = round((net / income) * 100, 1) if income > 0 else 0
    summary = {}
    mode_breakdown = {}
    cashflow_map = {}
    investment_txns = []
    for t in txns:
        mode = t.get("mode", "UPI") or "UPI"
        mode_breakdown[mode] = mode_breakdown.get(mode, 0) + 1
        day = t.get("date", "")
        if day not in cashflow_map:
            cashflow_map[day] = {"date": day, "income": 0, "spent": 0}
        if t["amount"] > 0:
            cashflow_map[day]["income"] += t["amount"]
        else:
            cashflow_map[day]["spent"] += abs(t["amount"])
            cat = t.get("category", "Others") or "Others"
            if cat not in summary:
                summary[cat] = {"total": 0, "count": 0, "avg": 0, "color": t["color"], "icon": t["icon"]}
            summary[cat]["total"] += abs(t["amount"])
            summary[cat]["count"] += 1
            if cat == "Investment":
                investment_txns.append({
                    "date": t["date"],
                    "name": t["description"],
                    "amount": round(abs(t["amount"]), 2),
                })
    for item in summary.values():
        item["total"] = round(item["total"], 2)
        item["avg"] = round(item["total"] / item["count"], 2) if item["count"] else 0
    amounts = [abs(t["amount"]) for t in debits]
    mean = (sum(amounts) / len(amounts)) if amounts else 0
    std = ((sum((a - mean) ** 2 for a in amounts) / len(amounts)) ** 0.5) if amounts else 0
    anomalies = [{
        "date": t["date"],
        "description": t["description"],
        "amount": t["amount"],
        "category": t["category"],
        "confidence": t["confidence"],
    } for t in debits if std and abs(abs(t["amount"]) - mean) > 2 * std]
    forecast = {
        cat: round(values["total"] * 1.08, 2)
        for cat, values in summary.items()
        if values["total"] > 0
    }
    cashflow = [cashflow_map[k] for k in sorted(cashflow_map)]
    investment_total = round(sum(t["amount"] for t in investment_txns), 2)
    top_cat = None
    if summary:
        top_cat = max(summary.items(), key=lambda item: item[1]["total"])
    insights = []
    if income > 0 or spent > 0:
        insights.append(f"You earned ₹{income:,.0f} and spent ₹{spent:,.0f} - savings rate: {rate}%.")
    if top_cat:
        cat_name, cat_stats = top_cat
        pct = round((cat_stats["total"] / spent) * 100) if spent else 0
        insights.append(f"{cat_name} is top expense at {pct}% (₹{cat_stats['total']:,.0f}).")
    if anomalies:
        insights.append(f"⚠ {len(anomalies)} unusual transaction(s) flagged.")
    if forecast:
        insights.append(f"Next month forecast: ₹{sum(forecast.values()):,.0f} projected spending.")
    return {
        "transactions": txns,
        "summary": summary,
        "income": income,
        "spent": spent,
        "net": net,
        "savings_rate": rate,
        "mode_breakdown": mode_breakdown,
        "anomalies": anomalies,
        "forecast": forecast,
        "bills": [],
        "investments": {
            "detected_txns": investment_txns,
            "total_invested": investment_total,
            "count": len(investment_txns),
        },
        "goal_projection": {"goals": [], "monthly_savings": net, "savings_rate": rate},
        "insights": insights,
        "cashflow": cashflow,
        "total_txns": len(txns),
        "account": {},
    }
def _resolve_report_scope(conn, user_id, report_id=None):
    if not report_id:
        return None, None
    report = conn.execute("SELECT * FROM reports WHERE id=? AND user_id=?", (report_id, user_id)).fetchone()
    if not report:
        return None, None
    report = dict(report)
    return report, (report.get("storage_name") or report.get("name"))
def init_db():
    conn = get_db(); cur = conn.cursor()
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS users(
            id TEXT PRIMARY KEY,
            username TEXT UNIQUE,
            display_name TEXT,
            password_hash TEXT,
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS transactions(
            id TEXT PRIMARY KEY, date TEXT, description TEXT, amount REAL,
            type TEXT, category TEXT, mode TEXT, is_card INTEGER,
            balance REAL, confidence REAL, source_file TEXT);
        CREATE TABLE IF NOT EXISTS goals(
            id TEXT PRIMARY KEY, name TEXT, target REAL, saved REAL,
            color TEXT, created_at TEXT, deadline TEXT, completed INTEGER DEFAULT 0);
        CREATE TABLE IF NOT EXISTS achieved_goals(
            id TEXT PRIMARY KEY, name TEXT, target REAL, color TEXT, achieved_at TEXT);
        CREATE TABLE IF NOT EXISTS reports(
            id TEXT PRIMARY KEY, name TEXT, filed_at TEXT,
            total_income REAL, total_debit REAL, tx_count INTEGER,
            report_excel_b64 TEXT, summary_json TEXT, period TEXT,
            account_holder TEXT);
    """)
    for sql in [
        "ALTER TABLE transactions ADD COLUMN user_id TEXT DEFAULT ''",
        "ALTER TABLE transactions ADD COLUMN confidence REAL DEFAULT 0",
        "ALTER TABLE transactions ADD COLUMN balance REAL DEFAULT 0",
        "ALTER TABLE goals ADD COLUMN user_id TEXT DEFAULT ''",
        "ALTER TABLE achieved_goals ADD COLUMN user_id TEXT DEFAULT ''",
        "ALTER TABLE reports ADD COLUMN user_id TEXT DEFAULT ''",
        "ALTER TABLE reports ADD COLUMN storage_name TEXT DEFAULT ''",
        "ALTER TABLE reports ADD COLUMN report_excel_b64 TEXT",
        "ALTER TABLE reports ADD COLUMN summary_json TEXT",
        "ALTER TABLE reports ADD COLUMN period TEXT",
        "ALTER TABLE reports ADD COLUMN account_holder TEXT",
    ]:
        try: cur.execute(sql)
        except: pass
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_username ON users(username)")
    conn.commit(); conn.close()
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False
    print("⚠️  openpyxl not found — Excel export disabled. pip install openpyxl")
try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False
try:
    import numpy as np
    NP_OK = True
except ImportError:
    NP_OK = False
@app.after_request
def add_cors(resp):
    resp.headers.update({
        "Access-Control-Allow-Origin":  "*",
        "Access-Control-Allow-Headers": "Content-Type",
        "Access-Control-Allow-Methods": "GET,POST,DELETE,OPTIONS",
    })
    return resp
@app.route("/", defaults={"path": ""}, methods=["OPTIONS"])
@app.route("/<path:path>", methods=["OPTIONS"])
def handle_options(path=""): return jsonify({}), 200
@app.route("/api/auth/me")
def auth_me():
    return jsonify(auth_response_payload())
@app.route("/api/auth/signup", methods=["POST"])
def auth_signup():
    data = request.get_json(silent=True) or {}
    display_name = (data.get("display_name") or "").strip()
    username = normalize_username(data.get("username"))
    password = data.get("password") or ""
    confirm_password = data.get("confirm_password") or ""
    if len(display_name) < 2:
        return jsonify({"error": "Display name must be at least 2 characters"}), 400
    if len(username) < 3:
        return jsonify({"error": "Username must be at least 3 characters"}), 400
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    if password != confirm_password:
        return jsonify({"error": "Passwords do not match"}), 400
    conn = get_db()
    try:
        existing = conn.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
        if existing:
            return jsonify({"error": "Username already exists"}), 409
        user_id = str(uuid.uuid4())
        conn.execute(
            "INSERT INTO users (id, username, display_name, password_hash, created_at) VALUES (?,?,?,?,?)",
            (user_id, username, display_name, generate_password_hash(password), datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
        session["user_id"] = user_id
        session["username"] = username
        session["display_name"] = display_name
        return jsonify(auth_response_payload()), 201
    finally:
        conn.close()
@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json(silent=True) or {}
    username = normalize_username(data.get("username"))
    password = data.get("password") or ""
    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400
    conn = get_db()
    try:
        user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        if not user or not check_password_hash(user["password_hash"], password):
            return jsonify({"error": "Invalid username or password"}), 401
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["display_name"] = user["display_name"] or user["username"]
        return jsonify(auth_response_payload())
    finally:
        conn.close()
@app.route("/api/auth/logout", methods=["POST"])
@login_required
def auth_logout():
    session.clear()
    return jsonify({"ok": True})
CAT_COLORS = {
    "Food": "#f59e0b", "Shopping": "#a855f7", "Entertainment": "#ef4444",
    "Utilities": "#3b82f6", "Health": "#10b981", "Transport": "#06b6d4",
    "Investment": "#7c3aed", "Transfer": "#f472b6", "Education": "#8b5cf6",
    "Income": "#22c55e", "Others": "#6b7280",
}
CAT_HEX = {k: v.lstrip("#") for k, v in CAT_COLORS.items()}
CATEGORY_KEYWORDS = {
    "Food":          ["swiggy","zomato", "SWIGGY","ZOMATO","mcdonald","kfc","pizza","restaurant","cafe","biryani","burger","bakery","sweets","food","meal","dhaba","canteen","hotel","bawa","la scents","sri amman","kannur cocktail","amman sweets"],
    "Shopping":      ["amazon","flipkart","myntra","ajio","nykaa","dmart","bigbazaar","mall","store","retail","clothes","shoes","electronics","mobile","vaishali metal","mangal","amazonpay","meesho","decathlon"],
    "Entertainment": ["netflix","prime","hotstar","spotify","youtube","apple","movie","pvr","inox","game","steam","angel one","angelone","tata play","tataplay","google play","bookmyshow","disney","supermoney"],
    "Utilities":     ["electricity","jio","airtel","vodafone","bsnl","wifi","internet","broadband","recharge","mobile bill","utility","jiofiber","jio postpaid","google india digital","gpay-utility","paytm","tneb","ptybl","tataplaybinge","apy","apyscf","pension","lombard","insurance","icici lom","myjio"],
    "Health":        ["gym","pharmacy","hospital","doctor","lab","clinic","medic","apollo","cult fit","1mg","diagnostic","wellness","health"],
    "Transport":     ["uber","ola","rapido","irctc","makemytrip","cleartrip","indigo","spicejet","bus","train","flight","oyo","petrol","fuel","toll","fastag","metro","taxi","cab","diesel"],
    "Investment":    ["zerodha","groww","angel one","angelone","upstox","mutual fund","sip","nps","demat","brokerage","iccl","brk","paying angel","angel one limited"],
    "Transfer":      ["transfer","neft","imps","rtgs","self","emi","credit card bill","card bill","loan","saravanan","chandrasekaran","kaasu","senthilkumar","senthil kumar"],
    "Education":     ["college","school","university","institute","coaching","tuition","fee","vmkv","engineering college","course","admission"],
    "Income":        ["salary","dividend","interest","credited","refund","cashback","bonus","received","payment from","deposit","bna seq","atm"],
    "Others":        [],
}
def ai_categorise(desc):
    d = (desc or "").lower()
    for cat, kws in CATEGORY_KEYWORDS.items():
        if cat != "Others" and any(kw in d for kw in kws):
            return cat
    return "Others"
def _mode_source(desc):
    d = (desc or "").upper()
    # Indian Bank rows include slash-separated "BRANCH : ..." metadata.
    # It can mention ATM SERVICE BRANCH even for UPI transactions.
    parts = [p.strip() for p in d.split("/")]
    return " / ".join(p for p in parts if p and not re.match(r"^BRANCH\s*:", p))
def detect_mode(desc):
    d = _mode_source(desc)
    if re.search(r"(^|[/\s])UPI([/\s]|$)", d) or "@" in d: return "UPI"
    if any(x in d for x in ("NEFT", "IMPS", "RTGS")):      return "Net Banking"
    if any(x in d for x in ("CARD", "POS", "ECOM", "MERCHNT", "TERMINAL ID")):
        return "Card"
    if any(x in d for x in ("BNA", "ATM ID", "ATM WDL", "ATM CASH", "CASH WITHDRAWAL")):
        return "ATM"
    return "Bank"
def is_card_txn(desc):
    return any(k in (desc or "").lower() for k in ["card","credit card","debit card","visa","mastercard","rupay","pos","swipe","ecom"])
def parse_inr(val):
    if not val: return 0.0
    s = re.sub(r"[^\d.]", "", str(val).strip().replace("INR","").replace(",",""))
    try: return float(s) if s else 0.0
    except: return 0.0
def _inr(v):
    if v is None: return None
    s = str(v).strip()
    if s in ("-"," - ","","None","Debits","Credits","Balance"): return None
    # Strip PNB-style " Cr." / " Dr." suffix from balance column
    s = re.sub(r"\s*(cr\.?|dr\.?)$", "", s, flags=re.IGNORECASE).strip()
    n = re.sub(r"[^\d.]", "", s)
    return float(n) if n else None
def _date_str(s):
    if not s: return None
    MO = {m:i+1 for i,m in enumerate(["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"])}
    t = str(s).strip()
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})$", t)
    if m:
        mo = MO.get(m.group(2).lower())
        if mo: return f"{m.group(3)}-{mo:02d}-{int(m.group(1)):02d}"
    if re.match(r"^\d{4}-\d{2}-\d{2}$", t): return t
    m2 = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", t)
    if m2: return f"{m2.group(3)}-{m2.group(2)}-{m2.group(1)}"
    # DD-MM-YYYY (Axis Bank and others)
    m3 = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", t)
    if m3: return f"{m3.group(3)}-{m3.group(2)}-{m3.group(1)}"
    return None
def _clean_desc(raw):
    s = str(raw or "").strip()
    if not s: return ""
    # Named transaction patterns → fixed labels
    if re.search(r"BNA\s*SEQ|TRAN\s*DATE", s, re.I):   return "ATM Cash Deposit"
    if re.search(r"ATM_AMC|BULK CHARGES", s, re.I):    return "Bank Charges"
    if re.search(r"SMS_CHGS|SMS CHRG", s, re.I):       return "SMS Charges"
    if re.search(r"CREDIT INTEREST|Int\.Pd", s, re.I): return "Credit Interest"
    if re.search(r"MIN BAL CHGS", s, re.I):            return "Min Balance Charges"
    if re.search(r"APYSCF|APY.NPST", s, re.I):         return "APY Pension Fund"
    if re.search(r"LOAN RECOVERY|Loan Recovery", s):   return "Loan Repayment"
    # Tokens to always discard regardless of position
    _SKIP = {
        "UPI","NEFT","IMPS","RTGS","ATM","DR","CR","IN","OUT",
        "MOB","SELFFT",                                    # Axis mobile/self-transfer
        "P2A","P2M","P2C","P2P","P2V","P2U",              # UPI transfer-type codes
        "UPI CO","COLLEC","PAYMEN","REMARK",               # Axis truncated labels
        "PHONE","UTIB","SBIN","HDFC","ICICI",              # bank short codes
    }
    for p in s.split("/")[1:]:
        p = p.strip()
        if not p: continue
        if p.upper() in _SKIP: continue
        if re.match(r"^XXXXX", p, re.I): continue
        if "@" in p: continue
        if re.match(r"^\d{4,}$", p): continue
        if "BRANCH" in p.upper(): continue
        if len(p) <= 2: continue
        if re.match(r"^(Punjab|Canara|Axis|HDFC|ICICI|SBI|State|Indian|Yes|Kotak|Federal|Airtel|Paytm)",
                    p, re.I): continue
        c = re.sub(r"\s+", " ", p).strip()
        if len(c) >= 3: return c[:80]
    return s[:200].strip()
def parse_xlsx_statement(raw_bytes_or_path):
    """Parse Indian Bank / generic .xlsx — returns (txns_list, meta_dict)"""
    if not XLSX_OK:
        raise ImportError("pip install openpyxl")
    if isinstance(raw_bytes_or_path, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes_or_path), data_only=True)
    else:
        wb = openpyxl.load_workbook(raw_bytes_or_path, data_only=True)
    ws = wb.active
    meta = {"holder":"","account":"","branch":"","ifsc":"","period":"","bank":"Indian Bank"}
    if ws.title and "axis" in ws.title.lower(): meta["bank"] = "Axis Bank"
    for row in ws.iter_rows(values_only=True):
        row = list(row) + [None]*17
        r0 = str(row[0] or "").strip()
        r1 = str(row[1] or "").strip()
        r3 = str(row[3] or "").strip()
        r4 = str(row[4] or "").strip() if row[4] else ""
        if "axis" in r0.lower() or "axis" in r1.lower(): meta["bank"] = "Axis Bank"
        # PNB meta
        if "account statement for account number" in r0.lower():
            meta["bank"] = "PNB"
            acc = re.sub(r"[^\d]", "", r0)
            if acc: meta["account"] = acc
        if r0.lower().startswith("customer name:") and r1:  meta["holder"] = r1
        if r0.lower().startswith("branch name:")   and r1:  meta["branch"] = r1
        if r0.lower().startswith("ifsc:")          and r1:  meta["ifsc"]   = r1
        if "statement period" in r0.lower():
            meta["period"] = re.sub(r"statement period[:\s]*", "", r0, flags=re.I).strip()
        # Generic / Axis meta
        if r1 == "Account Holder Name" and r4: meta["holder"]  = r4
        if r1 == "Account Number"       and r3: meta["account"] = r3
        if r1 == "Branch Name"          and r3: meta["branch"]  = r3
        if r1 == "IFSC"                 and r3: meta["ifsc"]    = r3
        if r1 and str(r1).startswith("For period"):
            meta["period"] = str(r1).replace("For period:","").strip()
    HEADER_SYNONYMS = {
        "date":    ["txn date","date","transaction date","value date","trans date","tran date"],
        "desc":    ["description","narration","particulars","transaction details","remarks"],
        "debit":   ["debit amount","debit","withdrawal","dr amount","dr amount","dr"],
        "credit":  ["credit amount","credit","deposit","cr amount","cr amount","cr"],
        "balance": ["balance","bal","closing balance","running balance"],
    }
    col_map = {}
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        matched = {}
        for role, synonyms in HEADER_SYNONYMS.items():
            for ci, cell in enumerate(cells):
                if cell in synonyms:
                    matched[role] = ci; break
        if "date" in matched and "desc" in matched and ("debit" in matched or "credit" in matched):
            col_map = matched; break
    if not col_map:
        col_map = {"date": 1, "desc": 2, "debit": 8, "credit": 10, "balance": 13}
    ci_date  = col_map.get("date",   1)
    ci_desc  = col_map.get("desc",   2)
    ci_debit = col_map.get("debit",  8)
    ci_cred  = col_map.get("credit", 10)
    ci_bal   = col_map.get("balance",13)
    txns, balance = [], 0.0
    for row in ws.iter_rows(values_only=True):
        row = list(row) + [None]*17
        raw_date = row[ci_date]
        if hasattr(raw_date, "strftime"):
            d = raw_date.strftime("%Y-%m-%d")
        else:
            ds = str(raw_date or "").strip()
            if not ds or ds in ("Date","Txn Date","Ending Balance","Total","") or "Indian Bank" in ds:
                continue
            d = _date_str(ds)
        if not d: continue
        raw_desc = str(row[ci_desc] or "").strip()
        if not raw_desc or raw_desc in ("Description","BRANCH","Narration","Particulars","PARTICULARS"): continue
        desc = _clean_desc(raw_desc)
        if not desc: continue
        deb = _inr(row[ci_debit]); cred = _inr(row[ci_cred]); bal = _inr(row[ci_bal])
        if deb is None and cred is None: continue
        amt = cred if cred is not None else -(deb or 0)
        if bal is not None: balance = bal
        else: balance += amt
        txns.append({"date":d,"description":desc,"amount":round(amt,2),
                     "balance":round(balance,2),"mode":detect_mode(raw_desc),"raw_desc":raw_desc})
    txns.sort(key=lambda r: r["date"])
    return txns, meta
def parse_csv_statement(fh_or_str):
    """Universal CSV parser — works for Indian Bank, HDFC, SBI, ICICI, Axis, Kotak, etc."""
    from parser import parse_statement
    if hasattr(fh_or_str, "read"):
        raw = fh_or_str.read()
    elif isinstance(fh_or_str, (bytes, bytearray)):
        raw = fh_or_str
    else:
        raw = str(fh_or_str).encode("utf-8")
    return parse_statement(raw)

def parse_pdf_statement(fp):
    """Parse PDF bank statement — extracts tables first, falls back to text."""
    from parser import parse_statement
    if not PDF_OK:
        raise ImportError("pip install pdfplumber")
    rows = []
    with pdfplumber.open(fp) as pdf:
        for page in pdf.pages:
            # Try structured table extraction first
            for table in (page.extract_tables() or []):
                for row in table:
                    if not row: continue
                    rows.append(",".join(str(c or "").replace(",", " ") for c in row))
            # Also grab raw text as fallback
            text = page.extract_text() or ""
            if text:
                rows.append(text)
    combined = "\n".join(rows)
    return parse_statement(combined.encode("utf-8"))
DB = {
    "transactions": [], "filed_reports": [],
    "clusters": {k:0 for k in CAT_COLORS},
    "summary": {"total_credit":0,"total_debit":0,"net_position":0,"weekly":[0,0,0,0],"liquidity_used_pct":0},
    "mode_counts": {"UPI":0,"Card":0,"Cash":0,"Bank":0,"Net Banking":0,"ATM":0},
    "card_transactions": [],
}
def reload_txns(conn, user_id):
    sync_storage(conn, user_id)
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM transactions WHERE user_id=? ORDER BY date ASC",
        (user_id,)
    ).fetchall()]
    DB["transactions"] = rows
    clusters = {k:0 for k in CAT_COLORS}
    mode_counts = {"UPI":0,"Card":0,"Cash":0,"Bank":0,"Net Banking":0,"ATM":0}
    total_credit = total_debit = 0
    weekly = [0,0,0,0]
    now = datetime.now()
    for t in rows:
        cat = t.get("category","Others")
        if cat in clusters: clusters[cat] = clusters.get(cat,0) + max(0, t.get("amount",0) if t.get("type")=="debit" else 0)
        m = t.get("mode","UPI")
        if m in mode_counts: mode_counts[m] += 1
        if t.get("type") == "credit": total_credit += t.get("amount",0)
        else: total_debit += t.get("amount",0)
        try:
            dt = datetime.strptime(t["date"], "%Y-%m-%d")
            delta = (now - dt).days
            if 0 <= delta < 7: weekly[3] += abs(t.get("amount",0))
            elif 7 <= delta < 14: weekly[2] += abs(t.get("amount",0))
            elif 14 <= delta < 21: weekly[1] += abs(t.get("amount",0))
            elif 21 <= delta < 28: weekly[0] += abs(t.get("amount",0))
        except: pass
    net = total_credit - total_debit
    liq = round(total_debit / total_credit * 100, 1) if total_credit > 0 else 0
    DB["clusters"] = clusters
    DB["mode_counts"] = mode_counts
    DB["summary"] = {"total_credit":total_credit,"total_debit":total_debit,
                     "net_position":net,"weekly":weekly,"liquidity_used_pct":min(100,liq)}
    DB["card_transactions"] = [t for t in rows if t.get("is_card") and t.get("type")=="debit"]
def summary_response():
    return {"clusters": DB["clusters"], "summary": DB["summary"], "mode_counts": DB["mode_counts"]}
def build_report_excel(ml_result, meta=None):
    """Build a styled per-category Excel report from ML result. Returns bytes."""
    if not XLSX_OK:
        raise ImportError("openpyxl required — pip install openpyxl")
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb       = openpyxl.Workbook()
    txns     = ml_result.get("transactions", [])
    summary  = ml_result.get("summary", {})
    account  = meta or ml_result.get("account", {})
    total_spent = ml_result.get("spent", 1) or 1
    def fill(hex6): return PatternFill("solid", fgColor=hex6.lstrip("#"))
    def thin():
        s = Side(style="thin", color="E2E8F0")
        return Border(left=s,right=s,top=s,bottom=s)
    def cw(ws, widths):
        for i,w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    WHITE="FFFFFF"; DARK="1E293B"; GREY="64748B"
    L_FILL = PatternFill("solid", fgColor="F8FAFC")
    W_FILL = PatternFill("solid", fgColor="FFFFFF")
    H_FILL = PatternFill("solid", fgColor="7C3AED")
    cats   = sorted(summary.items(), key=lambda x: x[1]["total"], reverse=True)
    ws = wb.active; ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    cw(ws, [2,24,14,16,14,16,2])
    ws.merge_cells("B1:F1"); ws.row_dimensions[1].height = 30
    c = ws["B1"]
    c.value = "FinUP. — Categorised Financial Report"
    c.font  = Font(name="Calibri", bold=True, size=16, color="7C3AED")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B2:F2"); ws.row_dimensions[2].height = 16
    c = ws["B2"]
    c.value = f"Account: {account.get('holder','')}  ·  {account.get('period','')}  ·  A/C: {account.get('account','')}"
    c.font  = Font(name="Calibri", size=10, color=GREY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 44
    kpis = [("B","Total Credits",ml_result.get("income",0),"D1FAE5","166534"),
            ("C","Total Debits", ml_result.get("spent",0), "FEE2E2","991B1B"),
            ("D","Net Savings",  ml_result.get("net",0),   "DBEAFE","1D4ED8"),
            ("E","Transactions", ml_result.get("total_txns",0),"EDE9FE","6D28D9")]
    for col, label, val, bg, fg in kpis:
        cell = ws[f"{col}4"]
        display = f"₹{abs(val):,.2f}" if isinstance(val, float) else str(val)
        cell.value = f"{label}\n{display}"
        cell.font  = Font(name="Calibri", bold=True, size=11, color=fg)
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin()
    ws.row_dimensions[6].height = 22
    for ci,h in enumerate(["Category","Transactions","Avg / Txn","% of Spend","Total Spent"], 2):
        c = ws.cell(row=6, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = H_FILL; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin()
    for ri, (cat, sv) in enumerate(cats, 7):
        ws.row_dimensions[ri].height = 20
        pct = round(sv["total"]/total_spent*100, 1)
        hex6 = CAT_HEX.get(cat, "6B7280")
        alt  = L_FILL if ri % 2 == 0 else W_FILL
        for ci, val in enumerate([cat, sv["count"], sv["avg"], f"{pct}%", sv["total"]], 2):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = alt; c.border = thin()
            if ci == 2:
                c.font = Font(name="Calibri", bold=True, size=10, color=hex6)
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif ci in (3,5):
                c.number_format = '₹#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.font = Font(name="Calibri", size=10, color=DARK, bold=(ci==5))
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Calibri", size=10, color=DARK)
    tr = 7 + len(cats)
    ws.row_dimensions[tr].height = 22
    tot_fill = PatternFill("solid", fgColor="EDE9FE")
    for ci, val in enumerate(["TOTAL", sum(v["count"] for _,v in cats), "", "", total_spent], 2):
        c = ws.cell(row=tr, column=ci, value=val)
        c.fill = tot_fill; c.border = thin()
        c.font = Font(name="Calibri", bold=True, size=10, color=DARK)
        if ci == 6: c.number_format = '₹#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
        else: c.alignment = Alignment(horizontal="center", vertical="center")
    ws_ins = wb.create_sheet("AI Insights")
    ws_ins.sheet_view.showGridLines = False
    cw(ws_ins, [2, 80, 2])
    ws_ins.merge_cells("B1:B1"); ws_ins.row_dimensions[1].height = 28
    c = ws_ins["B1"]; c.value = "AI-Generated Insights"
    c.font = Font(name="Calibri", bold=True, size=14, color="7C3AED")
    for ri, ins in enumerate(ml_result.get("insights", []), 3):
        ws_ins.row_dimensions[ri].height = 22
        c = ws_ins.cell(row=ri, column=2, value=f"• {ins}")
        c.font = Font(name="Calibri", size=11, color=DARK)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.fill = L_FILL if ri % 2 == 0 else W_FILL
    forecast = ml_result.get("forecast", {})
    if forecast:
        ws_fc = wb.create_sheet("Forecast")
        ws_fc.sheet_view.showGridLines = False
        cw(ws_fc, [2, 24, 16, 2])
        c = ws_fc["B1"]; c.value = "Next Month Spending Forecast"
        c.font = Font(name="Calibri", bold=True, size=13, color="7C3AED")
        for ci, h in enumerate(["Category", "Forecast (₹)"], 2):
            c = ws_fc.cell(row=3, column=ci, value=h)
            c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill = H_FILL; c.alignment = Alignment(horizontal="center")
        for ri, (cat, val) in enumerate(sorted(forecast.items(), key=lambda x:-x[1]), 4):
            hex6 = CAT_HEX.get(cat, "6B7280")
            c = ws_fc.cell(row=ri, column=2, value=cat)
            c.font = Font(name="Calibri", bold=True, size=10, color=hex6)
            c2 = ws_fc.cell(row=ri, column=3, value=val)
            c2.number_format = '₹#,##0.00'
            c2.font = Font(name="Calibri", size=10, color=DARK)
            c2.alignment = Alignment(horizontal="right")
    for cat, sv in cats:
        cat_txns = [t for t in txns if t.get("category") == cat]
        if not cat_txns: continue
        hex6  = CAT_HEX.get(cat, "6B7280")
        cname = re.sub(r'[\\/*?:\[\]]', '_', cat)[:28]
        ws2 = wb.create_sheet(title=cname)
        ws2.sheet_view.showGridLines = False
        cw(ws2, [2,14,42,14,12,14,14,2])
        CAT_FILL = PatternFill("solid", fgColor=hex6)
        ws2.merge_cells("B1:G1"); ws2.row_dimensions[1].height = 30
        c = ws2["B1"]
        c.value = f"{cat}  —  {len(cat_txns)} transactions  ·  Total: ₹{sv['total']:,.2f}"
        c.font  = Font(name="Calibri", bold=True, size=14, color=hex6)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.merge_cells("B2:G2"); ws2.row_dimensions[2].height = 16
        c = ws2["B2"]
        c.value = f"Period: {account.get('period','')}  ·  Account: {account.get('holder','')}  ·  Avg: ₹{sv['avg']:,.2f}"
        c.font  = Font(name="Calibri", size=10, color=GREY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[4].height = 20
        for ci, h in enumerate(["Date","Description","Mode","Confidence","Amount (₹)","Balance (₹)"], 2):
            c = ws2.cell(row=4, column=ci, value=h)
            c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill = CAT_FILL; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin()
        for ri, t in enumerate(cat_txns, 5):
            ws2.row_dimensions[ri].height = 18
            alt = L_FILL if ri%2==0 else W_FILL
            conf = t.get("confidence", 0)
            for ci, val in enumerate([t["date"], t["description"], t.get("mode","UPI"),
                                       f"{conf:.1f}%" if isinstance(conf,float) else f"{conf}%",
                                       t["amount"], t.get("balance",0)], 2):
                c = ws2.cell(row=ri, column=ci, value=val)
                c.fill = alt; c.border = thin()
                if ci == 2:   c.font = Font(name="Calibri",size=10,color=GREY); c.alignment=Alignment(horizontal="center")
                elif ci == 3: c.font = Font(name="Calibri",size=10,color=DARK); c.alignment=Alignment(horizontal="left",vertical="center")
                elif ci == 6:
                    c.number_format = '₹#,##0.00'; c.alignment=Alignment(horizontal="right")
                    c.font = Font(name="Calibri",size=10,bold=True,color="166534" if val>=0 else "991B1B")
                elif ci == 7:
                    c.number_format = '₹#,##0.00'; c.alignment=Alignment(horizontal="right")
                    c.font = Font(name="Calibri",size=10,color=GREY)
                else:
                    c.font = Font(name="Calibri",size=10,color=DARK); c.alignment=Alignment(horizontal="center")
        fr = 5 + len(cat_txns) + 1
        ws2.row_dimensions[fr].height = 22
        ws2.merge_cells(f"B{fr}:E{fr}")
        c  = ws2.cell(row=fr, column=2, value=f"Total — {cat}")
        c.font  = Font(name="Calibri",bold=True,size=11,color=hex6)
        c.fill  = PatternFill("solid",fgColor="F5F3FF"); c.alignment=Alignment(horizontal="right",vertical="center"); c.border=thin()
        c2 = ws2.cell(row=fr, column=6, value=sv["total"])
        c2.number_format = '₹#,##0.00'
        c2.font = Font(name="Calibri",bold=True,size=11,color=hex6)
        c2.fill = PatternFill("solid",fgColor="F5F3FF"); c2.alignment=Alignment(horizontal="right",vertical="center"); c2.border=thin()
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()
def simple_predict(txns):
    """Lightweight predict() compatible with ML.predict() output."""
    for t in txns:
        t["category"]   = ai_categorise(t.get("description",""))
        t["confidence"] = 85.0 if t["category"] != "Others" else 40.0
        t["color"]      = CAT_COLORS.get(t["category"],"#6b7280")
        t["icon"]       = "💳"
    credits = [t for t in txns if t["amount"] > 0]
    debits  = [t for t in txns if t["amount"] < 0]
    income  = sum(t["amount"] for t in credits)
    spent   = sum(abs(t["amount"]) for t in debits)
    net     = income - spent
    rate    = round(net/income*100,1) if income > 0 else 0
    summary = {}
    for t in debits:
        cat = t["category"]
        if cat not in summary:
            summary[cat] = {"total":0,"count":0,"avg":0,"color":t["color"],"icon":t["icon"]}
        summary[cat]["total"] += abs(t["amount"]); summary[cat]["count"] += 1
    for v in summary.values():
        v["avg"] = round(v["total"]/v["count"],2); v["total"] = round(v["total"],2)
    amounts = [abs(t["amount"]) for t in debits]
    mean = sum(amounts)/len(amounts) if amounts else 0
    std  = (sum((a-mean)**2 for a in amounts)/len(amounts))**0.5 if amounts else 0
    anomalies = [{"date":t["date"],"description":t["description"],"amount":t["amount"],
                  "category":t["category"],"confidence":t["confidence"]}
                 for t in debits if abs(abs(t["amount"])-mean) > 2*std]
    forecast = {cat: round(sv["total"]*1.08,2) for cat,sv in summary.items()}
    mode_breakdown = {}
    for t in txns: mode_breakdown[t.get("mode","UPI")] = mode_breakdown.get(t.get("mode","UPI"),0)+1
    cashflow_map = {}
    for t in txns:
        d = t["date"]
        if d not in cashflow_map: cashflow_map[d] = {"date":d,"income":0,"spent":0}
        if t["amount"]>0: cashflow_map[d]["income"] += t["amount"]
        else: cashflow_map[d]["spent"] += abs(t["amount"])
    cashflow = sorted(cashflow_map.values(), key=lambda x:x["date"])
    cs = sorted(summary.items(), key=lambda x:-x[1]["total"])
    insights = []
    if income > 0: insights.append(f"You earned ₹{income:,.0f} and spent ₹{spent:,.0f} — savings rate: {rate}%.")
    if cs: insights.append(f"{cs[0][0]} is top expense at {round(cs[0][1]['total']/spent*100) if spent else 0}% (₹{cs[0][1]['total']:,.0f}).")
    if anomalies: insights.append(f"⚠ {len(anomalies)} unusual transaction(s) flagged.")
    insights.append(f"Next month forecast: ₹{sum(forecast.values()):,.0f} projected spending.")
    return {
        "transactions": txns,"summary":summary,"income":round(income,2),"spent":round(spent,2),
        "net":round(net,2),"savings_rate":rate,"mode_breakdown":mode_breakdown,
        "anomalies":anomalies,"forecast":forecast,"bills":[],"investments":{"detected_txns":[],"total_invested":0,"count":0},
        "goal_projection":{"goals":[],"monthly_savings":round(net,2),"savings_rate":rate},
        "insights":insights,"cashflow":cashflow,"total_txns":len(txns),
    }
@app.route("/")
@app.route("/index.html")
def serve_index(): return send_from_directory(".", "index.html")
@app.route("/script.js")
def serve_script(): return send_from_directory(".", "script.js")
@app.route("/ml_engine.py")
def block_ml(): return "", 404
@app.route("/api/ml/upload", methods=["POST"])
@app.route("/upload_statement", methods=["POST"])
@login_required
def upload_statement():
    """Accept file upload, parse, run ML, store in SQLite, return full ML result."""
    user_id = current_user_id()
    txns, meta, raw, filename = [], {}, b"", "upload"
    storage_name = ""
    if "file" in request.files:
        f = request.files["file"]
        filename = clean_upload_name(f.filename or "upload")
        raw = f.read()
        storage_name = build_storage_name(user_id, filename)
        name = filename.lower()
        save_uploaded_statement(storage_name, raw)
        if name.endswith((".xlsx",".xls")):
            try:   txns, meta = parse_xlsx_statement(raw)
            except Exception as e: return jsonify({"error":f"Excel error: {e}"}), 400
        elif name.endswith((".csv",".txt")):
            txns = parse_csv_statement(raw)
            meta = {"holder":"Account Holder","bank":"Universal","period":"","account":""}
        elif name.endswith(".pdf"):
            tmp = upload_path(storage_name)
            try:   txns = parse_pdf_statement(tmp)
            except Exception as e: return jsonify({"error":f"PDF error: {e}"}), 400
            meta = {"holder":"Account Holder","bank":"Indian Bank","period":"","account":""}
        else:
            return jsonify({"error":"Upload .xlsx, .csv, or .pdf"}), 400
    else:
        return jsonify({"error":"No file provided"}), 400
    if not txns:
        return jsonify({"error":"No transactions found. Check your file format."}), 400
    try:
        if ML is not None:
            result = ML.predict(txns)
        else:
            result = simple_predict(txns)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error":f"ML error: {e}"}), 500
    result["account"] = {
        "holder":     meta.get("holder",""),
        "bank":       meta.get("bank","Indian Bank"),
        "period":     meta.get("period",""),
        "account_no": meta.get("account",""),
        "branch":     meta.get("branch",""),
        "ifsc":       meta.get("ifsc",""),
    }
    conn = get_db()
    conn.execute("DELETE FROM transactions WHERE user_id=? AND source_file=?", (user_id, storage_name))
    conn.execute("DELETE FROM reports WHERE user_id=? AND storage_name=?", (user_id, storage_name))
    conn.commit()
    for t in result.get("transactions", []):
        rid = str(uuid.uuid4())
        conn.execute(
            "INSERT OR IGNORE INTO transactions (id, date, description, amount, type, category, mode, is_card, balance, confidence, source_file, user_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (rid, t["date"], t["description"], abs(t["amount"]),
             "credit" if t["amount"]>0 else "debit",
             t.get("category","Others"), t.get("mode","UPI"),
             int(is_card_txn(t["description"])),
             t.get("balance",0), t.get("confidence",0), storage_name, user_id)
        )
    conn.commit()
    reload_txns(conn, user_id)
    report_excel_b64 = None
    try:
        if XLSX_OK:
            excel_bytes = build_report_excel(result, meta)
            report_excel_b64 = base64.b64encode(excel_bytes).decode("utf-8")
    except Exception as e:
        print(f"⚠️  Excel build error: {e}")
    report_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO reports (id, name, filed_at, total_income, total_debit, tx_count, report_excel_b64, summary_json, period, account_holder, user_id, storage_name) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        (report_id, filename, datetime.now().strftime("%Y-%m-%d"),
         result.get("income",0), result.get("spent",0), result.get("total_txns",0),
         report_excel_b64,
         json.dumps(result.get("summary",{})),
         meta.get("period",""), meta.get("holder",""), user_id, storage_name)
    )
    conn.commit()
    filed = [dict(r) for r in conn.execute("SELECT * FROM reports WHERE user_id=? ORDER BY filed_at DESC", (user_id,)).fetchall()]
    goal_payload = fetch_goal_payload(conn, user_id)
    conn.close()
    result["report_id"]      = report_id
    result["db_saved"]       = len(result.get("transactions",[]))
    result["filed_reports"]  = filed
    result["saved_goals"]    = goal_payload["goals"]
    result["achieved_goals"] = goal_payload["achieved_goals"]
    result["goal_stats"]     = goal_payload["stats"]
    return jsonify(result)
@app.route("/process_clusters", methods=["POST"])
@login_required
def process_clusters():
    user_id = current_user_id()
    data     = request.get_json(silent=True) or {}
    filename = clean_upload_name(data.get("filename",""))
    storage_name = data.get("storage_name", "").strip()
    if not storage_name:
        return jsonify({"error":"storage_name is required"}), 400
    fp       = upload_path(storage_name)
    if not os.path.exists(fp): return jsonify({"error":"File not found"}), 404
    ext = filename.lower().rsplit(".",1)[-1]
    try:
        if ext == "xlsx":   txns, meta = parse_xlsx_statement(fp)
        elif ext == "csv":
            with open(fp,"rb") as fh: txns = parse_csv_statement(fh.read()); meta={}
        elif ext == "pdf":  txns = parse_pdf_statement(fp); meta={}
        else: return jsonify({"error":f"Unsupported: .{ext}"}), 400
    except Exception as e: return jsonify({"error":str(e)}), 500
    if not txns: return jsonify({"error":"No transactions found"}), 422
    try:
        result = ML.predict(txns) if ML else simple_predict(txns)
    except Exception as e:
        return jsonify({"error":f"ML: {e}"}), 500
    conn = get_db()
    conn.execute("DELETE FROM transactions WHERE user_id=? AND source_file=?", (user_id, storage_name))
    conn.execute("DELETE FROM reports WHERE user_id=? AND storage_name=?", (user_id, storage_name))
    conn.commit()
    for t in result.get("transactions",[]):
        conn.execute("INSERT OR IGNORE INTO transactions (id, date, description, amount, type, category, mode, is_card, balance, confidence, source_file, user_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                     (str(uuid.uuid4()), t["date"], t["description"], abs(t["amount"]),
                      "credit" if t["amount"]>0 else "debit",
                      t.get("category","Others"), t.get("mode","UPI"),
                      int(is_card_txn(t["description"])), t.get("balance",0), t.get("confidence",0), storage_name, user_id))
    conn.commit(); reload_txns(conn, user_id)
    report_excel_b64 = None
    try:
        if XLSX_OK:
            excel_bytes = build_report_excel(result, meta)
            report_excel_b64 = base64.b64encode(excel_bytes).decode()
    except:
        pass
    conn.execute("INSERT INTO reports (id, name, filed_at, total_income, total_debit, tx_count, report_excel_b64, summary_json, period, account_holder, user_id, storage_name) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                 (str(uuid.uuid4()), filename, datetime.now().strftime("%Y-%m-%d"),
                  result.get("income",0), result.get("spent",0), result.get("total_txns",0),
                  report_excel_b64, json.dumps(result.get("summary",{})),
                  meta.get("period",""), meta.get("holder",""), user_id, storage_name))
    conn.commit()
    filed = [dict(r) for r in conn.execute("SELECT * FROM reports WHERE user_id=? ORDER BY filed_at DESC", (user_id,)).fetchall()]
    conn.close()
    return jsonify({"status":"success","transactions_found":len(txns),
                    "filed_reports":filed,**summary_response()})
@app.route("/api/report/download/<rid>")
@login_required
def download_report(rid):
    user_id = current_user_id()
    conn = get_db()
    row  = conn.execute("SELECT name, report_excel_b64 FROM reports WHERE id=? AND user_id=?", (rid, user_id)).fetchone()
    conn.close()
    if not row or not row["report_excel_b64"]:
        return jsonify({"error":"Report not found or no Excel stored"}), 404
    excel_bytes = base64.b64decode(row["report_excel_b64"])
    base_name   = os.path.splitext(row["name"] or "report")[0]
    return send_file(
        io.BytesIO(excel_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"FinUP_{base_name}_Categorised.xlsx"
    )
@app.route("/api/report/rebuild", methods=["POST"])
@login_required
def rebuild_report():
    """Build categorised Excel directly from posted ML result (browser fallback)."""
    data      = request.json or {}
    ml_result = data.get("ml_result", {})
    meta      = data.get("account", {})
    if not ml_result:
        return jsonify({"error":"No ml_result provided"}), 400
    try:
        excel_bytes = build_report_excel(ml_result, meta)
        return send_file(io.BytesIO(excel_bytes),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True,
                         download_name="FinUP_Categorised_Report.xlsx")
    except Exception as e:
        return jsonify({"error":str(e)}), 500
@app.route("/dashboard")
@login_required
def get_dashboard():
    user_id = current_user_id()
    conn = get_db(); reload_txns(conn, user_id); conn.close()
    return jsonify(summary_response())
@app.route("/transactions")
@login_required
def get_transactions():
    user_id = current_user_id()
    conn = get_db(); reload_txns(conn, user_id); conn.close()
    period = request.args.get("period","1M")
    days   = {"1M":30,"6M":180,"12M":365}.get(period,30)
    cutoff = (datetime.now()-timedelta(days=days)).strftime("%Y-%m-%d")
    filtered = [t for t in DB["transactions"] if (t.get("date","") or "") >= cutoff]
    return jsonify({"transactions":filtered[-50:], "mode_counts":DB["mode_counts"],
                    "card_transactions":[t for t in filtered if t.get("is_card") and t.get("type")=="debit"][-10:],
                    "total":len(filtered)})
@app.route("/api/bootstrap")
@login_required
def bootstrap_state():
    user_id = current_user_id()
    conn = get_db()
    try:
        sync_storage(conn, user_id)
        report_id = request.args.get("report_id", "").strip()
        scoped_report, scoped_name = _resolve_report_scope(conn, user_id, report_id)
        if report_id and not scoped_name:
            return jsonify({"error": "Report not found"}), 404
        if scoped_name:
            rows = [dict(r) for r in conn.execute(
                "SELECT * FROM transactions WHERE user_id=? AND source_file=? ORDER BY date ASC",
                (user_id, scoped_name)
            ).fetchall()]
        else:
            rows = [dict(r) for r in conn.execute("SELECT * FROM transactions WHERE user_id=? ORDER BY date ASC", (user_id,)).fetchall()]
        reports = [dict(r) for r in conn.execute("SELECT * FROM reports WHERE user_id=? ORDER BY filed_at DESC", (user_id,)).fetchall()]
        for report in reports:
            report.pop("report_excel_b64", None)
            report.pop("summary_json", None)
            report.pop("storage_name", None)
            report.pop("user_id", None)
        ml_result = build_persisted_ml_result(rows)
        goal_payload = fetch_goal_payload(conn, user_id)
        ml_result["active_report"] = {
            "id": scoped_report["id"],
            "name": scoped_report["name"],
            "period": scoped_report.get("period", ""),
            "filed_at": scoped_report.get("filed_at", ""),
        } if scoped_report else None
        ml_result["filed_reports"] = reports
        ml_result["saved_goals"] = goal_payload["goals"]
        ml_result["achieved_goals"] = goal_payload["achieved_goals"]
        ml_result["goal_stats"] = goal_payload["stats"]
        ml_result["user"] = current_user_profile()
        return jsonify(ml_result)
    except sqlite3.Error as exc:
        empty = build_persisted_ml_result([])
        empty["filed_reports"] = []
        empty["saved_goals"] = []
        empty["achieved_goals"] = []
        empty["goal_stats"] = {"monthly_savings": 0, "savings_rate": 0, "achieved_count": 0, "active_count": 0, "total_target": 0, "total_saved": 0}
        empty["active_report"] = None
        empty["error"] = str(exc)
        return jsonify(empty), 200
    finally:
        conn.close()
@app.route("/api/trends")
@login_required
def get_trends():
    """Return extended trend data: monthly cashflow, category trend, savings trend."""
    user_id = current_user_id()
    conn = get_db()
    report_id = request.args.get("report_id", "").strip()
    _, scoped_name = _resolve_report_scope(conn, user_id, report_id)
    if report_id and not scoped_name:
        conn.close()
        return jsonify({"error": "Report not found", "monthly": []}), 404
    if scoped_name:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM transactions WHERE user_id=? AND source_file=? ORDER BY date ASC",
            (user_id, scoped_name)
        ).fetchall()]
    else:
        rows = [dict(r) for r in conn.execute("SELECT * FROM transactions WHERE user_id=? ORDER BY date ASC", (user_id,)).fetchall()]
    conn.close()
    monthly: dict = {}
    for t in rows:
        try:
            dt = datetime.strptime(t["date"][:10],"%Y-%m-%d")
            key = dt.strftime("%Y-%m")
        except: continue
        if key not in monthly:
            monthly[key] = {"month":key,"income":0,"spent":0,"net":0,"by_cat":{}}
        if t.get("type")=="credit": monthly[key]["income"] += t.get("amount",0)
        else:
            monthly[key]["spent"] += t.get("amount",0)
            cat = t.get("category","Others")
            monthly[key]["by_cat"][cat] = monthly[key]["by_cat"].get(cat,0) + t.get("amount",0)
    for m in monthly.values():
        m["net"] = round(m["income"]-m["spent"],2)
        m["income"] = round(m["income"],2); m["spent"] = round(m["spent"],2)
    return jsonify({"monthly": sorted(monthly.values(), key=lambda x:x["month"])})
@app.route("/goals", methods=["GET"])
@login_required
def get_goals():
    user_id = current_user_id()
    conn  = get_db()
    payload = fetch_goal_payload(conn, user_id)
    conn.close()
    return jsonify(payload)
@app.route("/goals", methods=["POST"])
@login_required
def add_goal():
    user_id = current_user_id()
    b = request.json or {}
    target = float(b["target"])
    saved = float(b.get("saved",0))
    completed = 1 if target and saved >= target else 0
    g = {"id":str(uuid.uuid4()),"name":b["name"],"target":target,
         "saved":saved,"color":b.get("color","#7c3aed"),
         "created_at":datetime.now().strftime("%Y-%m-%d"),"deadline":b.get("deadline","Open"),"completed":completed}
    conn = get_db()
    conn.execute("INSERT INTO goals (id, name, target, saved, color, created_at, deadline, completed, user_id) VALUES (?,?,?,?,?,?,?,?,?)",
                 (g["id"],g["name"],g["target"],g["saved"],g["color"],g["created_at"],g["deadline"],g["completed"], user_id))
    if completed and not conn.execute("SELECT id FROM achieved_goals WHERE id=? AND user_id=?", (g["id"], user_id)).fetchone():
        conn.execute("INSERT INTO achieved_goals (id, name, target, color, achieved_at, user_id) VALUES (?,?,?,?,?,?)",
                     (g["id"], g["name"], g["target"], g["color"], datetime.now().strftime("%Y-%m-%d"), user_id))
    conn.commit(); conn.close()
    return jsonify({**g, "percent": 100.0 if completed else round(saved / target * 100, 1) if target else 0})
@app.route("/goals/<gid>", methods=["POST"])
@login_required
def update_goal(gid):
    user_id = current_user_id()
    saved = float(request.json.get("saved",0))
    conn  = get_db()
    row   = conn.execute("SELECT * FROM goals WHERE id=? AND user_id=?", (gid, user_id)).fetchone()
    just_completed = False
    if row:
        completed = 1 if saved >= row["target"] else 0
        just_completed = completed==1 and not bool(row["completed"])
        conn.execute("UPDATE goals SET saved=?, completed=? WHERE id=? AND user_id=?", (saved,completed,gid,user_id))
        if just_completed and not conn.execute("SELECT id FROM achieved_goals WHERE id=? AND user_id=?", (gid, user_id)).fetchone():
            conn.execute("INSERT INTO achieved_goals (id, name, target, color, achieved_at, user_id) VALUES (?,?,?,?,?,?)",
                         (gid,row["name"],row["target"],row["color"],datetime.now().strftime("%Y-%m-%d"), user_id))
    conn.commit()
    updated = _goal_to_dict(conn.execute("SELECT * FROM goals WHERE id=? AND user_id=?", (gid, user_id)).fetchone())
    conn.close()
    return jsonify({"status":"updated","goal":updated,"just_completed":just_completed})
@app.route("/goals/<gid>", methods=["DELETE"])
@login_required
def delete_goal(gid):
    user_id = current_user_id()
    conn = get_db()
    conn.execute("DELETE FROM goals WHERE id=? AND user_id=?", (gid, user_id))
    conn.execute("DELETE FROM achieved_goals WHERE id=? AND user_id=?", (gid, user_id))
    conn.commit()
    payload = fetch_goal_payload(conn, user_id)
    conn.close()
    return jsonify(payload)
STOCK_BASE = {"NIFTY":24142.70,"SENSEX":72831.94,"GOLD":14110.00,"SILVER":260.00}
@app.route("/investments/prices")
def get_prices():
    prices = {}
    for sym, base in STOCK_BASE.items():
        p = round(base*(1+random.uniform(-0.5,0.5)/100), 2)
        prices[sym] = {"price":p,"change":round(p-base,2),"change_pct":round((p-base)/base*100,2)}
    return jsonify({"prices":prices,"updated_at":datetime.now().isoformat()})
@app.route("/reports")
@login_required
def get_reports():
    user_id = current_user_id()
    conn    = get_db(); reload_txns(conn, user_id)
    reports = [dict(r) for r in conn.execute("SELECT * FROM reports WHERE user_id=? ORDER BY filed_at DESC", (user_id,)).fetchall()]
    conn.close()
    for r in reports:
        r.pop("report_excel_b64", None)
        r.pop("summary_json", None)
        r.pop("storage_name", None)
        r.pop("user_id", None)
    return jsonify({"filed_reports":reports,**summary_response()})
@app.route("/reports/<rid>", methods=["DELETE"])
@login_required
def delete_report(rid):
    user_id = current_user_id()
    conn = get_db()
    row  = conn.execute("SELECT name, storage_name FROM reports WHERE id=? AND user_id=?", (rid, user_id)).fetchone()
    if row:
        storage_name = row["storage_name"] or row["name"]
        conn.execute("DELETE FROM transactions WHERE user_id=? AND source_file=?", (user_id, storage_name))
        try:
            os.remove(upload_path(storage_name))
        except FileNotFoundError:
            pass
    conn.execute("DELETE FROM reports WHERE id=? AND user_id=?", (rid, user_id)); conn.commit()
    reload_txns(conn, user_id)
    reports = [dict(r) for r in conn.execute("SELECT * FROM reports WHERE user_id=? ORDER BY filed_at DESC", (user_id,)).fetchall()]
    for r in reports:
        r.pop("report_excel_b64",None); r.pop("summary_json",None); r.pop("storage_name", None); r.pop("user_id", None)
    conn.close()
    return jsonify({"filed_reports":reports,**summary_response()})
@app.route("/api/ml/train", methods=["POST"])
def ml_train():
    if ML is None: return jsonify({"error":"ML engine not loaded"}), 503
    try:
        r = ML.train()
        return jsonify({"ok":True,"result":{k:str(v) for k,v in r.items()}})
    except Exception as e:
        return jsonify({"error":str(e)}), 500
@app.route("/api/db-status")
def db_status():
    return jsonify({"connected":True,"message":"SQLite connected ✅","engine":"sqlite"})
@app.route("/upload_statement_only", methods=["POST"])
@login_required
def upload_only():
    if "file" not in request.files: return jsonify({"error":"No file"}), 400
    f = request.files["file"]
    filename = clean_upload_name(f.filename)
    storage_name = build_storage_name(current_user_id(), filename)
    save_uploaded_statement(storage_name, f.read())
    return jsonify({"filename": filename, "storage_name": storage_name})
@app.route("/pay_bills", methods=["POST"])
def pay_bills():
    return jsonify({"status":"redirect","redirect_url":"https://paydesk.example.com/checkout",
                    "bills":[{"name":"Credit Card (HDFC)","amount":124000,"due":"Next billing cycle"},
                             {"name":"Internet/Wifi","amount":5500,"due":"Next billing cycle"}]})
if __name__ == "__main__":
    init_db()
    conn = get_db()
    sync_storage(conn)
    conn.close()
    if ML is not None:
        mp = os.path.join(_ML_DIR, "category_model.pkl")
        if not os.path.exists(mp):
            print("🤖 Training ML model…"); ML.train(); print("✅ ML model ready")
    print("  🚀 FinUP.  →  http://127.0.0.1:5000/index.html")
    print("  🧠 ML: TF-IDF + Logistic Regression categories")
    app.run(debug=True, port=5000)